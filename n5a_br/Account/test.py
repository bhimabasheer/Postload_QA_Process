import openpyxl
from openpyxl.styles import numbers, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from glob import glob
from os import getcwd
import os

_date = "m/d/yyyy"
_num = "0.0000"
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for  file in glob(r"Z:\Dataload\Beema\Post_load_QA\n5a_br\Account\*.xlsx"):
    print(file)
    
    wb = load_workbook(file)

    ws = wb['Balance']
   
    for row_num, row in enumerate(ws.iter_rows(), start=1):
        for cell_num, cell in enumerate(row, start=1):
            if ws.cell(row=1, column=cell_num).value == "RowNum":
                continue
            
  
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

  
            if cell.data_type == 'n':
                cell.number_format = _num

            elif cell.data_type == 'd':
                cell.number_format = _date

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)

        col_letter = get_column_letter(column[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(r'Z:\Dataload\Beema\Post_load_QA\n5a_br\Account\sample_Sample_Data_Report.xlsx')
    wb.close()
