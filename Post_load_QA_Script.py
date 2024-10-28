import os
from urllib.parse import quote_plus
import openpyxl
from openpyxl.styles import numbers, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from glob import glob
from os import getcwd
from sqlalchemy import create_engine,text,MetaData,Table
from configparser import RawConfigParser
import argparse
from datetime import datetime

# Initialize input for procedure and dynamic code running

_date = "m/d/yyyy"
_num = "0.0000"
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

parser = argparse.ArgumentParser()
parser.add_argument('-c', '--ClientID',action='store',help="ClientID")
parser.add_argument('-f', '--Filercode',action='store',help="FilerCode")
args = parser.parse_args()

# Getting userid and password from properties file
path=f"{os.getcwd()}\\credentials.properties"

def credentials(filepath,USER):
    config = RawConfigParser()
    config.read(filepath)
    return {"UID" : config.get(f"{USER}","UID"),"PWD": config.get(f"{USER}","PWD")}
 
USER="BEEMA"
    
UID = credentials(path,USER).get("UID")

PWD = credentials(path,USER).get("PWD")

# Establish connection string
connection_string = (
'Driver={ODBC Driver 17 For SQL Server};'
'SERVER=standoutresearch.cmgm5ibackyh.us-east-2.rds.amazonaws.com;'
'Database=Data_Load;'
f'UID={UID};'
f'PWD={PWD};'
'Trusted_Connection=no;'
)

connection_uri = f"mssql+pyodbc:///?odbc_connect={quote_plus(connection_string)}"

engine = create_engine(connection_uri, fast_executemany=True)

connection=engine.connect()

metadata = MetaData() 

Client_ID = args.ClientID

Filer_code = args.Filercode

table_name = 'Postload_QA_Details'

table = Table(table_name, metadata, autoload_with=engine)

# Execute the stored procedure 1
Get_postload_QA = text(f"EXEC Get_postload_QA @ClientId = '{Client_ID}', @FilerCode = '{Filer_code}'")

PostLoad_ReportGen = text("EXEC PostLoad_ReportGen @IssueID = :param3, @ReportId = :param4")

# Define report id from procedure along with issues in QA table
dic={'A_Acc_Num_Null':[1,'Account'],'A_Acc_Num_Exp':[2,'Account'],
     'Mul_Acc_Name':[3,'Account'],'B_Acc_Num_Null':[4,'Balance'],
     'B_Acc_Num_Exp':[5,'Balance'],'B_CUSIP_Exp':[6,'Balance'],
     'B_ISIN_Exp':[7,'Balance'],'B_SEDOL_Exp':[8,'Balance'],
     'B_TICKER_Exp':[9,'Balance'],'B_AsofDate_Null':[10,'Balance'],
     'B_Blank_Security':[11,'Balance'],'B_Acc_Recon':[12,'Balance'],
     'T_Acc_Num_Null':[13,'Trade'],'T_Acc_Num_Exp':[14,'Trade'],
     'T_CUSIP_Exp':[15,'Trade'],'T_ISIN_Exp':[16,'Trade'],
     'T_SEDOL_Exp':[17,'Trade'],'T_TICKER_Exp':[18,'Trade'],
     'T_Trade_Date_Null':[19,'Trade'],'T_Blank_Security':[20,'Trade'],
     'T_Tran_Type_Null':[21,'Trade'],'T_Remapped_Trantype_Null':[22,'Trade'],
     'T_Acc_Recon':[23,'Trade'],'T_Duplicate_TranId':[24,'Trade']
    }

     

# Set folder access path
base_directory = r'D:\Beema\Post_load_QA'

try:
    # Return output from procedure 
    issueids = connection.execute(Get_postload_QA)
    # Get issue id
    isid=issueids.first()[0]
    connection.commit()


    
    # iterate report id ,subfolder names and issue column names
    for column, value in dic.items():       
        repid = value[0]
        Subfolder = value[1]
        # check issue present or not
        query = table.select().where((table.c[column] == 1) & (table.c['IssueID'] == isid))
    
        try:
            result = connection.execute(query).fetchall()         
  

            if result:  # Check if there are issues present
                print(Client_ID, Filer_code, isid, column, repid)

                # Specify the directory path dynamically
                directory_path = os.path.join(base_directory, f'{Client_ID}_{Filer_code}', f'{Subfolder}')

                # Ensure the specified directory exists
                if not os.path.exists(directory_path):
                    os.makedirs(directory_path)

                worksheet_name = f'{Subfolder}'

                excel_file_path = os.path.join(directory_path, f'{column}_{Client_ID}_{Filer_code}.xlsx')

                # Assign input for procedure 2
                Params_PostLoad_ReportGen = {'param3': isid, 'param4': repid}

                # Execute the second procedure for report generation
                report_data = connection.execute(PostLoad_ReportGen, Params_PostLoad_ReportGen)
                
                # Fetch results and headers
                rows = [list(row) for row in report_data]
                header = list(report_data.keys())

                # Create a new Excel workbook
                workbook = openpyxl.Workbook()
                worksheet = workbook.active               

                # Write header to the worksheet
                worksheet.title = worksheet_name
               
                
                worksheet.append(header)
                
                
                 
                #print(worksheet)

                # Write rows to the worksheet
                for row_data in rows:
                    #cleaned_row = ["" if cell is None else cell for cell in row_data]
                    try:                       
                        worksheet.append(row_data)
                    except Exception as e:
                        print(f"Error processing row: {e}")
                        print(f"Problematic row_data: {row_data}")              

                workbook.save(excel_file_path)
                workbook.close()

                wb = load_workbook(excel_file_path)

                ws = wb[f'{Subfolder}']
   
                for row_num, row in enumerate(ws.iter_rows(), start=1):
                    for cell_num, cell in enumerate(row, start=1):
                        if ws.cell(row=1, column=cell_num).value in("RowNum","Trans_ID","Accounts_ID"):
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            continue
                        
            
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
    
            
                        if cell.data_type == 'n':
                            cell.number_format = _num

                        elif cell.data_type == 'd':
                            cell.number_format = _date

                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 5)

                    col_letter = get_column_letter(column[0].column)
                    ws.column_dimensions[col_letter].width = adjusted_width          
                    wb.save(excel_file_path)
                    wb.close()      
                print("Data saved in path: " + excel_file_path)

        except Exception as e:
            print(f"Error processing column '{column}': {e}")

except Exception as e:
    print(f"Error processing : {e}")

finally:
    connection.close()
    engine.dispose()        
               